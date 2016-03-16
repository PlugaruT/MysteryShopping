from string import punctuation

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from .models import Evaluation
from .project_statuses import ProjectStatus


class EvaluationSpreadsheet:

    def __init__(self, evaluation=None):
        self.evaluation = evaluation if evaluation else None
        self.header_font_style = Font(name='Arial', size=18, bold=True)
        self.sub_header_font_style = Font(name='Arial', size=16, bold=True)
        self.default_font_style = Font(name='Arial', size=14)
        self.default_alignment = Alignment(vertical='center')
        self.header_height = 26
        self.sub_header_height = 23
        self.default_height_singleline = 22.5
        self.default_height_multiline = 21.5
        self.max_characters_line = 130 # 90
        self.divide = False
        self.big_merged_cell = 0
        self.small_merged_cell = 0
        self.evaluation_xlsx = Workbook()
        self.sheet = self.evaluation_xlsx.active
        self.column, self.row = 1, 1
        self.header_fill = PatternFill(fill_type='lightTrellis', start_color='FFFF0000', end_color='FFFFFF00')
        self.sub_header_fills = {0: PatternFill(fill_type='lightTrellis', start_color='FFFFF000', end_color='FFFFFFF0'),
                                 1: PatternFill(fill_type='lightTrellis', start_color='F1AB0340', end_color='FF234100'),
                                 2: PatternFill(fill_type='lightTrellis', start_color='12345678', end_color='87654321'),
                                 3: PatternFill(fill_type='lightTrellis', start_color='11111111', end_color='FFFFFFFF')}

    def generate_spreadsheet(self):
        if self.evaluation.status in ProjectStatus.EDITABLE_STATUSES:
            self.big_merged_cell = 9
            self.small_merged_cell = 0
        else:
            self.divide = True
            self.big_merged_cell = 7
            self.small_merged_cell = 2

        self.write_header('Script')
        self.write_sub_header('Title', height=self.sub_header_height)
        self.write_default(self.evaluation.questionnaire_script.title)
        self.write_sub_header('Description', height=self.sub_header_height)
        self.write_default(self.evaluation.questionnaire_script.description)

        self.row += 2

        self.write_header('Questionnaire', self.evaluation.questionnaire.score)
        for block in self.evaluation.questionnaire.blocks.all():
            self.iterate_block(block)

        return self.evaluation_xlsx

    def write_header(self, text, score=0):
        if self.divide and score:
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column,
                                   end_column=self.column+self.big_merged_cell)
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column+self.big_merged_cell+1,
                                   end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column)
            cell_to_write.font = self.header_font_style
            cell_to_write.fill = self.header_fill
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = text

            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column+self.big_merged_cell+1)
            cell_to_write.font = self.header_font_style
            cell_to_write.fill = self.header_fill
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = 'Score: {}'.format(str(score))
        else:
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column,
                                   end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column)
            cell_to_write.font = self.header_font_style
            cell_to_write.fill = self.header_fill
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = text

        self.sheet.row_dimensions[self.row].height = self.header_height
        self.row += 1

    def write_sub_header(self, text, height, score=0, fill=3):
        if self.divide and score:
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column,
                                   end_column=self.column+self.big_merged_cell)
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column+self.big_merged_cell+1,
                                   end_column=self.column+(self.big_merged_cell+self.small_merged_cell))

            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column)
            cell_to_write.fill = self.sub_header_fills[fill]
            cell_to_write.font = self.sub_header_font_style
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = text

            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column+self.big_merged_cell+1)
            cell_to_write.fill = self.sub_header_fills[fill]
            cell_to_write.font = self.sub_header_font_style
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = 'Score: {}'.format(str(score))
        else:
            self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                                   start_column=self.column,
                                   end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
            cell_to_write = self.sheet.cell(row=self.row,
                                            column=self.column)
            cell_to_write.font = self.sub_header_font_style
            cell_to_write.fill = self.sub_header_fills[fill]
            cell_to_write.alignment = self.default_alignment
            cell_to_write.value = text

        self.sheet.row_dimensions[self.row].height = height
        self.row += 1

    def write_default(self, text):
        self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                               start_column=self.column,
                               end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
        cell_to_write = self.sheet.cell(row=self.row, column=self.column)
        cell_to_write.alignment = Alignment(wrap_text=True, shrink_to_fit=True, vertical='center')
        cell_to_write.value = text
        cell_to_write.font = self.default_font_style
        rows_to_write = self.calculate_height(cell_to_write.value)
        if rows_to_write == 0:
            height = self.default_height_singleline
        else:
            height = self.default_height_multiline * rows_to_write
        self.sheet.row_dimensions[self.row].height = height
        self.row += 1

    def calculate_height(self, text):
        return int(len(text) / self.max_characters_line)

    def write_question(self, question):
        # TODO: check question type
        # Add a ':' at the end if question doesn't end with a punctuation sign
        if question.question_body.strip()[-1] not in punctuation:
            question_body = '{}:'.format(question.question_body)
        else:
            question_body = question.question_body
        self.write_default(text=question_body)
        if question.type in {'s', 'm'}:
            for question_choice in question.question_choices.all():
                check = question_choice.id in question.answer_choices if question.answer_choices else False
                self.write_question_choice(question_choice, question.type, check=check)
        else:
            self.write_question_choice_text()

    def write_question_choice(self, question_choice, type, check=False):
        type_of_shape = {'m': {True:  u"\u25A3",
                               False: u"\u25A1"},
                         's': {True:  u"\u25CF",
                               False: u"\u25CB"}}
        cell_to_write = self.sheet.cell(row=self.row, column=self.column)
        cell_to_write.font = self.default_font_style
        cell_to_write.alignment = Alignment(horizontal='center', vertical='center')
        cell_to_write.value = type_of_shape[type][check and self.divide]

        self.sheet.merge_cells(start_row=self.row, end_row=self.row,
                               start_column=self.column+1,
                               end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
        cell_to_write = self.sheet.cell(row=self.row, column=self.column+1)
        cell_to_write.alignment = Alignment(wrap_text=True, shrink_to_fit=True, )
        cell_to_write.value = question_choice.text
        cell_to_write.font = self.default_font_style
        cell_to_write.alignment = Alignment(vertical='center')
        rows_to_write = self.calculate_height(cell_to_write.value)
        if rows_to_write == 0:
            height = self.default_height_singleline
        else:
            height = self.default_height_multiline * rows_to_write
        self.sheet.row_dimensions[self.row].height = height
        self.row += 1

    def write_question_choice_text(self):
        # if self.divide:
        #     pass
        # else:
        self.sheet.merge_cells(start_row=self.row, end_row=self.row + 5,
                                   start_column=self.column,
                                   end_column=self.column+(self.big_merged_cell+self.small_merged_cell))
        self.row += 6

    def iterate_block(self, block):
        block_score = round(block.score / block.weight * 100, 2) if block.weight and block.score else 0
        self.write_sub_header(text=block.title, score=block_score, fill=block.level, height=self.sub_header_height*1.3)
        for question in block.questions.all():
            self.write_question(question)
        for sub_block in block.children.all():
            self.iterate_block(sub_block)
